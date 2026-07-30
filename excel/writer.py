from openpyxl import load_workbook
from datetime import datetime


COMMON_ROW = {
    "uptime": 3,

    "management_cpu": 5,
    "dataplane_cpu": 6,

    "total_memory": 9,
    "available_memory": 10,
    "used_memory": 11,

    "thermal": 24,
    "fan_tray": 25,
    "fans": 26,
    "power": 27,
    "power_supplies": 28,
    "log": 29,

    "uplink": 30,

    # Interface
    "ethernet1_1": 31,
    "ethernet1_9": 31,
    "ethernet1_10": 32,
    "ethernet1_19": 33,
    "ethernet1_20": 34,

    "ha_status": 35,
}


ROW_FW = {
    "dev_md3": 14,
    "none": 15,
    "dev_md5": 16,
    "dev_md6": 17,
    "tmpfs_dev_shm": 18,
    "dev_md8": 19,
    "tmpfs_run_dpdk": 20,
    "cgroup_root": 21,
    "tmpfs_ssl_private": 22,
}


ROW_RJ = {
    "dev_md3": 14,
    "none": 15,
    "dev_md5": 16,
    "dev_md6": 17,
    "tmpfs_dev_shm": 18,
    "tmpfs_run_dpdk": 19,
    "dev_md8": 20,
    "cgroup_root": 21,
    "tmpfs_ssl_private": 22,
}


def write_excel(excel_file, all_results):

    wb = load_workbook(excel_file)

    today = datetime.now().date()

    for sheet_name, result in all_results.items():

        print(result)

        ws = wb[sheet_name]

        target_col = None

        for col in range(3, ws.max_column + 1):

            value = ws.cell(row=2, column=col).value

            if isinstance(value, datetime):

                if value.date() == today:
                    target_col = col
                    break

        if target_col is None:
            print(f"Tanggal tidak ditemukan pada sheet {sheet_name}")
            continue

        row_map = COMMON_ROW.copy()

        if sheet_name in ["FW01", "FW02"]:
            row_map.update(ROW_FW)
        else:
            row_map.update(ROW_RJ)

        print(f"\n===== {sheet_name} =====")

        for key, value in result.items():

            if key not in row_map:
                continue

            cell = ws.cell(
                row=row_map[key],
                column=target_col
            )

            if key == "uptime":
                cell.number_format = "@"
                cell.value = str(value)
            else:
                cell.value = value

        print(f"{sheet_name} selesai")

    output = excel_file.replace(".xlsx", "_RESULT.xlsx")

    wb.save(output)

    print(f"\nSaved : {output}")